"""API auxiliar: exportar un reporte mensual en Excel (3 hojas estilizadas)."""
from __future__ import annotations

import datetime as _dt
import io
import re as _re

from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from ..db import get_db
from ..services import finanzas
from .helpers import household_actual

router = APIRouter(prefix="/api", tags=["api"])

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

MESES = ["", "enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
         "agosto", "septiembre", "octubre", "noviembre", "diciembre"]


def _fname(nombre: str | None, year: int, month: int, ext: str) -> str:
    """Nombre de archivo ASCII-seguro (los headers HTTP solo aceptan latin-1)."""
    slug = _re.sub(r"[^A-Za-z0-9]+", "_",
                   (nombre or "app").encode("ascii", "ignore").decode()).strip("_") or "familia"
    return f"finanzas-{slug}-{year}-{month:02d}.{ext}"


def _periodo(year: int | None, month: int | None, today: _dt.date) -> tuple[int, int]:
    """Año/mes saneados a un rango válido (evita 500 por fechas imposibles)."""
    return min(max(year or today.year, 2000), 2100), min(max(month or today.month, 1), 12)

# --- Paleta y estilos reutilizables ---------------------------------------
CLP = '"$"#,##0'
FECHA = "DD/MM/YYYY"
_AZUL = "1F2A52"
_AZUL2 = "2E3A6E"
_GRIS = "EEF1FA"
_VERDE = "1B8A4B"
_ROJO = "C0392B"
_MUTED = "6B7280"

_F_TITULO = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
_F_SUB = Font(name="Calibri", size=10, color="DDE3FF")
_F_HEAD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_F_BOLD = Font(name="Calibri", size=11, bold=True)
_F_KPI = Font(name="Calibri", size=14, bold=True)
_F_MUTED = Font(name="Calibri", size=10, color=_MUTED)
_FILL_TIT = PatternFill("solid", fgColor=_AZUL)
_FILL_HEAD = PatternFill("solid", fgColor=_AZUL2)
_FILL_GRIS = PatternFill("solid", fgColor=_GRIS)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_thin = Side(style="thin", color="D5DAE8")
_BORDE = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _anchos(ws: Worksheet, anchos: dict[str, int]) -> None:
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w


def _titulo(ws: Worksheet, ncols: int, titulo: str, subtitulo: str) -> None:
    fin = chr(ord("A") + ncols - 1)
    ws.merge_cells(f"A1:{fin}1")
    ws.merge_cells(f"A2:{fin}2")
    c = ws["A1"]; c.value = titulo; c.font = _F_TITULO; c.fill = _FILL_TIT; c.alignment = _LEFT
    s = ws["A2"]; s.value = subtitulo; s.font = _F_SUB; s.fill = _FILL_TIT; s.alignment = _LEFT
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 16


def _fila_head(ws: Worksheet, fila: int, headers: list[str]) -> None:
    for i, h in enumerate(headers):
        c = ws.cell(row=fila, column=i + 1, value=h)
        c.font = _F_HEAD; c.fill = _FILL_HEAD; c.alignment = _CENTER; c.border = _BORDE


@router.get("/export.xlsx")
def export_excel(request: Request, year: int | None = None, month: int | None = None,
                 db: Session = Depends(get_db)):
    household = household_actual(request, db)
    today = _dt.date.today()
    year, month = _periodo(year, month, today)

    wb = Workbook()
    if not household:
        ws = wb.active
        ws.title = "Sin datos"
        ws["A1"] = "Inicia sesión para exportar los datos de tu familia."
    else:
        r = finanzas.resumen_mes(db, household, year, month)
        txs = finanzas.transacciones_mes(db, household, year, month)
        bills = finanzas.bills_pendientes(db, household)
        sub = f"{household.name} · {MESES[month].capitalize()} {year}"
        _hoja_resumen(wb.active, household, sub, r)
        _hoja_movimientos(wb.create_sheet("Movimientos"), sub, txs)
        _hoja_cuentas(wb.create_sheet("Cuentas por pagar"), sub, bills, today)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = _fname(household.name if household else None, year, month, "xlsx")
    return StreamingResponse(buf, media_type=XLSX_MIME,
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ===========================================================================
# Export a PDF (reporte mensual imprimible)
# ===========================================================================
def _latin(s) -> str:
    """fpdf2 con fuentes core solo soporta latin-1: limpiamos emojis/símbolos raros."""
    return str(s or "").encode("latin-1", "replace").decode("latin-1")


def _clp_txt(n) -> str:
    return "$" + f"{int(n or 0):,}".replace(",", ".")


@router.get("/export.pdf")
def export_pdf(request: Request, year: int | None = None, month: int | None = None,
               db: Session = Depends(get_db)):
    from fpdf import FPDF

    household = household_actual(request, db)
    today = _dt.date.today()
    year, month = _periodo(year, month, today)

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    W = pdf.epw  # ancho útil

    if not household:
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, _latin("Inicia sesion para exportar tu reporte."))
        out = bytes(pdf.output())
        return StreamingResponse(io.BytesIO(out), media_type="application/pdf")

    r = finanzas.resumen_mes(db, household, year, month)
    txs = finanzas.transacciones_mes(db, household, year, month)
    bills = finanzas.bills_pendientes(db, household)

    # Encabezado
    pdf.set_fill_color(31, 42, 82)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 13, _latin(f"  Finanzas Familia"), fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11)
    pdf.set_fill_color(46, 58, 110)
    pdf.cell(0, 8, _latin(f"  {household.name} - {MESES[month].capitalize()} {year}"),
             fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(20, 20, 20)
    pdf.ln(5)

    # KPIs (2 columnas)
    kpis = [("Ingresos", r["ingresos"]), ("Gastos", r["gastos"]),
            ("Cuentas por pagar", r["bills_pendientes"]), ("Disponible", r["disponible"]),
            ("Gasto hormiga", r["hormigas"])]
    pdf.set_font("Helvetica", "", 11)
    cw = W / 2
    for i, (label, val) in enumerate(kpis):
        pdf.set_font("Helvetica", "", 10); pdf.set_text_color(110, 114, 128)
        pdf.cell(cw * 0.55, 7, _latin(label), border=0)
        pdf.set_font("Helvetica", "B", 11); pdf.set_text_color(20, 20, 20)
        nl = (i % 2 == 1) or (i == len(kpis) - 1)
        pdf.cell(cw * 0.45, 7, _clp_txt(val), border=0,
                 new_x="LMARGIN" if nl else "RIGHT", new_y="NEXT" if nl else "TOP")
    pdf.ln(4)

    def _tabla(titulo, headers, filas, anchos):
        pdf.set_font("Helvetica", "B", 12); pdf.set_text_color(20, 20, 20)
        pdf.cell(0, 8, _latin(titulo), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "B", 9); pdf.set_fill_color(31, 42, 82); pdf.set_text_color(255, 255, 255)
        for h, w in zip(headers, anchos):
            pdf.cell(w, 7, _latin(h), border=0, fill=True, align="C")
        pdf.ln(7)
        pdf.set_text_color(20, 20, 20); pdf.set_font("Helvetica", "", 9)
        pdf.set_fill_color(238, 241, 250)
        fill = False
        for fila in filas:
            for j, val in enumerate(fila):
                align = "R" if j == len(fila) - 1 else "L"
                pdf.cell(anchos[j], 6, _latin(val), border="B", fill=fill, align=align)
            pdf.ln(6)
            fill = not fill

    # Gastos por categoría
    if r["por_categoria"]:
        filas = [[f"{c['emoji']} {c['nombre']}".strip(), _clp_txt(c["total"]),
                  f"{round(c['total']*100/(r['gastos'] or 1))}%"] for c in r["por_categoria"]]
        filas.append(["TOTAL", _clp_txt(r["gastos"]), ""])
        _tabla("Gastos por categoria", ["Categoria", "Monto", "%"], filas, [W*0.5, W*0.3, W*0.2])
        pdf.ln(3)

    # Cuentas por pagar
    if bills:
        filas = [[b.due_date.strftime("%d/%m/%Y"), b.label, _clp_txt(b.amount or 0)] for b in bills]
        _tabla("Cuentas por pagar", ["Vence", "Cuenta", "Monto"], filas, [W*0.25, W*0.5, W*0.25])
        pdf.ln(3)

    # Movimientos
    filas = [[t.occurred_at.strftime("%d/%m"),
              "Gasto" if t.kind == "expense" else "Ingreso",
              (t.category.name if t.category else ""),
              (t.user.name if t.user else ""),
              _clp_txt(t.amount)] for t in txs]
    if filas:
        _tabla("Movimientos", ["Fecha", "Tipo", "Categoria", "Persona", "Monto"],
               filas, [W*0.13, W*0.15, W*0.32, W*0.25, W*0.15])

    out = bytes(pdf.output())
    fname = _fname(household.name, year, month, "pdf")
    return StreamingResponse(io.BytesIO(out), media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ---------------------------------------------------------------------------
# Hoja 1: Resumen
# ---------------------------------------------------------------------------
def _hoja_resumen(ws: Worksheet, household, sub: str, r: dict) -> None:
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    _anchos(ws, {"A": 26, "B": 16, "C": 14, "D": 4, "E": 22, "F": 16})
    _titulo(ws, 6, "📊 Finanzas Familia", sub)

    # KPIs (fila 4-5)
    kpis = [("Ingresos", r["ingresos"], _VERDE), ("Gastos", r["gastos"], _ROJO),
            ("Cuentas por pagar", r["bills_pendientes"], _MUTED),
            ("Disponible", r["disponible"], _VERDE if r["disponible"] >= 0 else _ROJO),
            ("🐜 Gasto hormiga", r["hormigas"], "B7791F")]
    fila = 4
    for i, (label, val, color) in enumerate(kpis):
        col = 1 + i
        lc = ws.cell(row=fila, column=col, value=label)
        lc.font = _F_MUTED; lc.alignment = _CENTER
        vc = ws.cell(row=fila + 1, column=col, value=val)
        vc.number_format = CLP; vc.font = Font(size=13, bold=True, color=color); vc.alignment = _CENTER
        vc.fill = _FILL_GRIS; lc.fill = _FILL_GRIS
        lc.border = _BORDE; vc.border = _BORDE
    ws.row_dimensions[fila + 1].height = 22

    # Tabla: gastos por categoría (desde fila 8)
    f = 8
    ws.cell(row=f, column=1, value="Gastos por categoría").font = _F_BOLD
    f += 1
    _fila_head(ws, f, ["Categoría", "Monto", "% del gasto"])
    total_gastos = r["gastos"] or 1
    f += 1
    for cat in r["por_categoria"]:
        nombre = f"{cat['emoji']} {cat['nombre']}".strip()
        ws.cell(row=f, column=1, value=nombre).border = _BORDE
        mc = ws.cell(row=f, column=2, value=cat["total"]); mc.number_format = CLP; mc.alignment = _RIGHT; mc.border = _BORDE
        pc = ws.cell(row=f, column=3, value=cat["total"] / total_gastos); pc.number_format = "0.0%"; pc.alignment = _RIGHT; pc.border = _BORDE
        f += 1
    tc = ws.cell(row=f, column=1, value="TOTAL"); tc.font = _F_BOLD; tc.border = _BORDE
    tt = ws.cell(row=f, column=2, value=r["gastos"]); tt.number_format = CLP; tt.font = _F_BOLD; tt.alignment = _RIGHT; tt.border = _BORDE
    ws.cell(row=f, column=3).border = _BORDE

    # Tabla: gasto por persona (columnas E-F, desde fila 8)
    fp = 8
    ws.cell(row=fp, column=5, value="Gasto por persona").font = _F_BOLD
    fp += 1
    for ci, h in ((5, "Persona"), (6, "Monto")):
        hc = ws.cell(row=fp, column=ci, value=h)
        hc.font = _F_HEAD; hc.fill = _FILL_HEAD; hc.alignment = _CENTER; hc.border = _BORDE
    fp += 1
    for p in r["por_persona"]:
        ws.cell(row=fp, column=5, value=p["nombre"]).border = _BORDE
        mc = ws.cell(row=fp, column=6, value=p["total"]); mc.number_format = CLP; mc.alignment = _RIGHT; mc.border = _BORDE
        fp += 1
    if not r["por_persona"]:
        ws.cell(row=fp, column=5, value="Sin gastos").font = _F_MUTED


# ---------------------------------------------------------------------------
# Hoja 2: Movimientos
# ---------------------------------------------------------------------------
def _hoja_movimientos(ws: Worksheet, sub: str, txs: list) -> None:
    ws.sheet_view.showGridLines = False
    _anchos(ws, {"A": 12, "B": 10, "C": 22, "D": 34, "E": 18, "F": 11, "G": 16})
    _titulo(ws, 7, "Movimientos", sub)
    head = 4
    _fila_head(ws, head, ["Fecha", "Tipo", "Categoría", "Descripción", "Persona", "Origen", "Monto"])

    fuente = {"ia": "🤖 IA", "whatsapp": "💬 WhatsApp", "manual": "✍️ Manual", "statement": "📄 Cartola"}
    f = head + 1
    total_gasto = total_ingreso = 0
    for t in txs:
        es_gasto = t.kind == "expense"
        if es_gasto:
            total_gasto += t.amount
        else:
            total_ingreso += t.amount
        ws.cell(row=f, column=1, value=t.occurred_at).number_format = FECHA
        ws.cell(row=f, column=2, value="Gasto" if es_gasto else "Ingreso")
        ws.cell(row=f, column=3, value=f"{t.category.emoji if t.category and t.category.emoji else ''} {t.category.name if t.category else ''}".strip())
        ws.cell(row=f, column=4, value=t.description or t.raw_text or "")
        ws.cell(row=f, column=5, value=t.user.name if t.user else "")
        ws.cell(row=f, column=6, value=fuente.get(t.source, t.source or ""))
        mc = ws.cell(row=f, column=7, value=t.amount)
        mc.number_format = CLP; mc.alignment = _RIGHT
        mc.font = Font(color=_ROJO if es_gasto else _VERDE, bold=True)
        for col in range(1, 8):
            cell = ws.cell(row=f, column=col)
            cell.border = _BORDE
            if f % 2 == 0:
                cell.fill = _FILL_GRIS
        f += 1

    if not txs:
        ws.cell(row=f, column=1, value="Sin movimientos este mes.").font = _F_MUTED
        f += 1

    # Totales
    f += 1
    ws.cell(row=f, column=6, value="Total ingresos").font = _F_BOLD
    ti = ws.cell(row=f, column=7, value=total_ingreso); ti.number_format = CLP; ti.font = Font(bold=True, color=_VERDE); ti.alignment = _RIGHT
    ws.cell(row=f + 1, column=6, value="Total gastos").font = _F_BOLD
    tg = ws.cell(row=f + 1, column=7, value=total_gasto); tg.number_format = CLP; tg.font = Font(bold=True, color=_ROJO); tg.alignment = _RIGHT

    ws.freeze_panes = f"A{head + 1}"
    if txs:
        ws.auto_filter.ref = f"A{head}:G{head + len(txs)}"


# ---------------------------------------------------------------------------
# Hoja 3: Cuentas por pagar
# ---------------------------------------------------------------------------
def _hoja_cuentas(ws: Worksheet, sub: str, bills: list, hoy: _dt.date) -> None:
    ws.sheet_view.showGridLines = False
    _anchos(ws, {"A": 14, "B": 30, "C": 16, "D": 16, "E": 16})
    _titulo(ws, 5, "⏰ Cuentas por pagar (pendientes)", sub)
    head = 4
    _fila_head(ws, head, ["Vence", "Cuenta", "Monto", "Estado", "Aviso (días antes)"])

    f = head + 1
    total = 0
    for b in bills:
        total += b.amount or 0
        dias = (b.due_date - hoy).days
        estado = "Vencida" if dias < 0 else ("Vence hoy" if dias == 0 else f"En {dias} día{'s' if dias != 1 else ''}")
        ws.cell(row=f, column=1, value=b.due_date).number_format = FECHA
        ws.cell(row=f, column=2, value=b.label)
        mc = ws.cell(row=f, column=3, value=b.amount or 0); mc.number_format = CLP; mc.alignment = _RIGHT
        ec = ws.cell(row=f, column=4, value=estado); ec.alignment = _CENTER
        if dias < 0:
            ec.font = Font(color=_ROJO, bold=True)
        elif dias <= 3:
            ec.font = Font(color="B7791F", bold=True)
        ws.cell(row=f, column=5, value=b.notify_days_before).alignment = _CENTER
        for col in range(1, 6):
            ws.cell(row=f, column=col).border = _BORDE
        f += 1

    if not bills:
        ws.cell(row=f, column=1, value="Nada pendiente 🎉").font = _F_MUTED
    else:
        ws.cell(row=f, column=2, value="TOTAL").font = _F_BOLD
        tc = ws.cell(row=f, column=3, value=total); tc.number_format = CLP; tc.font = _F_BOLD; tc.alignment = _RIGHT
    ws.freeze_panes = f"A{head + 1}"
